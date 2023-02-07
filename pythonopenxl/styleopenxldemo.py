from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title="Hcl"
#sheet["A1"].value=10
#sheet["B2"].value="test"
#sheet.cell(row=3,column=3).value="Hcl data"
"""
j=0
for i in range(10,60,10):
   j+=1
   sheet.cell(row=j,column=1).value=i

for row in sheet.iter_rows(min_row=1,max_row=5,max_col=2,min_col=2):
    for r in row:
        r.value=1"""
k=100
for col in sheet.iter_rows(min_row=1,max_row=1,max_col=5,min_col=1):
    for r in col:
        r.value=k
        k+=100
wb.save("c:/hcl2/demoopenxlwrite.xlsx")
